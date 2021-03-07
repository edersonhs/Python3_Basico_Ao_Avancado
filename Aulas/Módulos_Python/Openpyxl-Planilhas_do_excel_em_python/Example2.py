import openpyxl

pedidos = openpyxl.load_workbook('pedidos.xlsx')   # Carregando a planilha na variável
nome_planilhas = pedidos.sheetnames   # Vai retornar o nome das planilhas que o arquivo contem
# print(nome_planilhas)   # ['Página1']    sÓ tem uma pagina no arquivo

planilha1 = pedidos['Página1']   # Passando a pagina 1 para outra variável

# Alterando Célula (A planilha original não será alterada, uma nova devera ser salva)
# planilha1['B3'].value = 2200   # Coluna B linha 3 receberá 2200
for linha in range(5, 16):   # Adicionando conteúdo da linha 5 a linha 15
    planilha1.cell(linha, 1).value = 'TESTE'   # cell é  uma outra forma de acessar o conteúdo da coluna
    planilha1.cell(linha, 2).value = 'TESTE'
    planilha1.cell(linha, 3).value = 'TESTE'

pedidos.save('Nova_Planilha.xlsx')   # Salvando a nova planilha
