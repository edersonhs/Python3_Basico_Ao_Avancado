import openpyxl
from random import uniform

# Criando uma planilha pelo python
planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)   # Criando a primeira planinha
planilha.create_sheet('Planilha2', 1)   # Criando a segunda planilha, que fica na segunda pag

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 11):   # Inserindo o conteúdo da planilha1
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(1, 11):   # Inserindo o conteúdo da planilha2
    planilha2.cell(linha, 1).value = f'Luiz {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 2).value = f'Otávio {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 3).value = f'Joãozinho {linha} {round(uniform(10, 100), 2)}'

planilha.save('nova_planilha.xlsx')   # Salvando a planilha
